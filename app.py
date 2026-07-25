from flask import Flask, render_template, request, redirect, url_for, session, send_file, flash
from flask import send_from_directory, Response
import requests
import xml.etree.ElementTree as ET
from openpyxl.drawing.image import Image as XLImage
import openpyxl
import re
import os
import hashlib
from openpyxl import Workbook
import urllib.request
import time
from urllib.parse import urlparse
from local_config import ADD_GAME_PASSWORD, BGG_HEADERS, SECRET_KEY

app = Flask(__name__)

@app.route('/favicon.ico')
def favicon():
    return send_from_directory('static', 'favicon.ico')

app.secret_key = SECRET_KEY

EXCEL_FILE = "boardgames.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        sheet = wb.active
        sheet.append(['Название игры', 'ID на BGG', 'Мин. игроков', 'Макс. игроков', 
                      'Время игры (мин)', 'Рейтинг', 'Категории', 'Добавил пользователь', 'Картинка', 'Избранное'])
        sheet.row_dimensions[1].height = 100
        wb.save(EXCEL_FILE)

init_excel()

def download_and_insert_image(sheet, row_num, image_url, game_name, game_id):
    """Скачивает картинку в высоком качестве и вставляет в Excel"""
    try:
        from PIL import Image as PILImage
        import io
        from requests.adapters import HTTPAdapter
        from urllib3.util.retry import Retry

        static_img_dir = 'static/images'
        if not os.path.exists(static_img_dir):
            os.makedirs(static_img_dir)
        
        # Очищаем имя файла
        safe_name = re.sub(r'[\\/*?:"<>|\.]', '_', game_name)
        safe_name = re.sub(r'\s+', '_', safe_name)
        if len(safe_name) > 50:
            safe_name = safe_name[:50]
        
        img_filename = f'{safe_name}_{game_id}.jpg'
        img_path = os.path.join(static_img_dir, img_filename)
        
        # Сессия с автоматическими повторными попытками
        sess = requests.Session()
        retry_strategy = Retry(
            total=3,
            backoff_factor=2,
            status_forcelist=[429, 500, 502, 503, 504],
        )
        adapter = HTTPAdapter(max_retries=retry_strategy)
        sess.mount("https://", adapter)
        sess.mount("http://", adapter)
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        response = sess.get(
            image_url, 
            headers=headers, 
            timeout=(10, 60),
            stream=True
        )
        response.raise_for_status()
        
        content = b""
        for chunk in response.iter_content(chunk_size=8192):
            if chunk:
                content += chunk
        
        if not content:
            print(f"Пустой ответ от сервера для {image_url}")
            return None
        
        img_pil = PILImage.open(io.BytesIO(content))
        
        if img_pil.mode in ('RGBA', 'P', 'LA'):
            background = PILImage.new('RGB', img_pil.size, (255, 255, 255))
            if img_pil.mode == 'P':
                img_pil = img_pil.convert('RGBA')
            background.paste(
                img_pil, 
                mask=img_pil.split()[-1] if img_pil.mode in ('RGBA', 'LA') else None
            )
            img_pil = background
        elif img_pil.mode != 'RGB':
            img_pil = img_pil.convert('RGB')
        
        img_pil.save(img_path, 'JPEG', quality=95, optimize=True)
        print(f"Картинка сохранена: {img_path} | Размер: {img_pil.size}")
        
        xl_img = XLImage(img_path)
        xl_img.width = 80
        xl_img.height = 80
        sheet.add_image(xl_img, f'I{row_num}')
        sheet.row_dimensions[row_num].height = 90
        sheet.column_dimensions['I'].width = 15
        
        return f'images/{img_filename}'
        
    except requests.exceptions.Timeout:
        print(f"Таймаут при скачивании картинки: {image_url}")
        return None
    except requests.exceptions.ConnectionError as e:
        print(f"Ошибка подключения при скачивании картинки: {e}")
        return None
    except requests.exceptions.RequestException as e:
        print(f"Ошибка запроса: {e}")
        return None
    except Exception as e:
        print(f"Неожиданная ошибка при загрузке картинки: {e}")
        import traceback
        traceback.print_exc()
        return None


def contains_cyrillic(text):
    return bool(re.search('[\u0400-\u04FF]', text))


def get_all_categories():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    categories = set()
    for row in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        if row[6]:
            for category in row[6].split("; "):
                if category and category.strip():
                    categories.add(category.strip())
    return sorted(list(categories))


@app.route('/')
def index():
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    games_count = sum(1 for row in sheet.iter_rows(min_row=2, values_only=True) if row[0])
    return render_template('index.html', games_count=games_count)


@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)


@app.route('/proxy_image')
def proxy_image():
    """Прокси для загрузки картинок с BGG через сервер"""
    image_url = request.args.get('url')
    
    if not image_url:
        return '', 404
    
    # Проверяем что URL с BGG (безопасность)
    allowed_domains = ['geekdo-images.com', 'boardgamegeek.com', 'cf.geekdo-images.com']
    parsed = urlparse(image_url)
    
    if not any(domain in parsed.netloc for domain in allowed_domains):
        return 'Forbidden', 403
    
    # Кеш в папке static/cache
    cache_dir = 'static/cache'
    os.makedirs(cache_dir, exist_ok=True)
    
    # Имя файла кеша — хеш от URL
    url_hash = hashlib.md5(image_url.encode()).hexdigest()
    cache_path = os.path.join(cache_dir, f'{url_hash}.jpg')
    
    # Если уже скачали — отдаём из кеша
    if os.path.exists(cache_path):
        return send_file(cache_path, mimetype='image/jpeg')
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        response = requests.get(image_url, headers=headers, timeout=10)
        response.raise_for_status()
        
        # Сохраняем в кеш
        with open(cache_path, 'wb') as f:
            f.write(response.content)
        
        return send_file(cache_path, mimetype='image/jpeg')
    
    except requests.exceptions.RequestException as e:
        print(f"Ошибка прокси: {e}")
        return '', 404


def get_games_list():
    """Возвращает коллекцию в едином формате для каталога и фильтров."""
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    games_list = []
    import glob
    
    for row_idx in range(2, sheet.max_row + 1):
        name = sheet.cell(row=row_idx, column=1).value
        if name:
            game_id = sheet.cell(row=row_idx, column=2).value
            min_players = sheet.cell(row=row_idx, column=3).value
            max_players = sheet.cell(row=row_idx, column=4).value
            playing_time = sheet.cell(row=row_idx, column=5).value
            rating = sheet.cell(row=row_idx, column=6).value
            categories = sheet.cell(row=row_idx, column=7).value
            added_by = sheet.cell(row=row_idx, column=8).value
            favorite = sheet.cell(row=row_idx, column=10).value
            
            image_filename = None
            
            safe_name = re.sub(r'[\\/*?:"<>|\.]', '_', name)
            safe_name = re.sub(r'\s+', '_', safe_name)
            if len(safe_name) > 50:
                safe_name = safe_name[:50]
            
            image_pattern = f'static/images/{safe_name}_{game_id}.jpg'
            
            if os.path.exists(image_pattern):
                image_filename = f'images/{safe_name}_{game_id}.jpg'
            else:
                import glob
                matching_files = glob.glob(f'static/images/{safe_name}_*.jpg')
                if matching_files:
                    image_filename = matching_files[0].replace('static/', '').replace('\\', '/')
            
            games_list.append({
                'name': name,
                'id': game_id,
                'min_players': min_players,
                'max_players': max_players,
                'playing_time': playing_time,
                'rating': rating,
                'categories': categories,
                'added_by': added_by,
                'image_path': image_filename,
                'favorite': str(favorite).strip() if favorite else ''
            })
    
    workbook.close()
    return games_list


@app.route('/games')
def games():
    return render_template('games.html', games=get_games_list())


@app.route('/export_excel')
def export_excel():
    if os.path.exists(EXCEL_FILE):
        return send_file(EXCEL_FILE, as_attachment=True, download_name='boardgames.xlsx')
    else:
        flash('Файл не найден', 'error')
        return redirect(url_for('index'))


@app.route('/add_game', methods=['GET', 'POST'])
def add_game():
    if session.get('password_ok'):
        return render_template('add_game.html')
    
    if request.method == 'POST':
        password = request.form.get('password')
        if password == ADD_GAME_PASSWORD:
            session['password_ok'] = True
            flash('Пароль верный! Теперь можно добавлять игры', 'success')
            return render_template('add_game.html')
        else:
            flash('Неверный пароль!', 'error')
            return render_template('password_check.html')
    
    return render_template('password_check.html')


@app.route('/search_bgg', methods=['POST'])
def search_bgg():
    search_text = request.form.get('game_name')
    if search_text:
        search_query = search_text.replace(" ", "%20")
        search_url = f"https://boardgamegeek.com/xmlapi2/search?query={search_query}&type=boardgame"
        
        headers = {
            **BGG_HEADERS,
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        }
        
        response = requests.get(search_url, headers=headers)
        
        if response.status_code == 200:
            root = ET.fromstring(response.text)
            results = []
            
            items = root.findall(".//item")[:10]
            
            for item in items:
                name_elem = item.find(".//name")
                if name_elem is not None:
                    name = name_elem.get('value')
                    objectid = item.get('id')
                    
                    thumbnail_url = None
                    try:
                        details_url = f"https://boardgamegeek.com/xmlapi2/thing?id={objectid}"
                        details_response = requests.get(details_url, headers=headers, timeout=10)
                        
                        print(f"Запрос картинки для {name} (ID: {objectid}), статус: {details_response.status_code}")
                        
                        if details_response.status_code == 200:
                            details_root = ET.fromstring(details_response.text)
                            game = details_root.find(".//item")
                            if game is not None:
                                thumbnail_elem = game.find("thumbnail")
                                if thumbnail_elem is not None and thumbnail_elem.text:
                                    thumbnail_url = thumbnail_elem.text
                                    if thumbnail_url.startswith('//'):
                                        thumbnail_url = 'https:' + thumbnail_url
                                    print(f"Найдена картинка: {thumbnail_url}")
                                else:
                                    print(f"Нет thumbnail для {name}")
                            else:
                                print(f"Игра {objectid} не найдена в детальном API")
                        else:
                            print(f"Ошибка {details_response.status_code} при запросе деталей {objectid}")
                        
                        time.sleep(0.5)
                        
                    except Exception as e:
                        print(f"Ошибка при получении картинки для {name}: {e}")
                        thumbnail_url = None
                    
                    results.append({
                        'name': name,
                        'id': objectid,
                        'thumbnail': thumbnail_url
                    })
            
            return render_template('search_results.html', results=results, search_text=search_text)
        else:
            flash(f'Ошибка при поиске на BGG (код {response.status_code})', 'error')
            return redirect(url_for('add_game'))
    
    return redirect(url_for('add_game'))


@app.route('/add_game_by_id', methods=['POST'])
def add_game_by_id():
    game_id = request.form.get('game_id')
    
    if not game_id or not game_id.isdigit():
        flash('Пожалуйста, введите корректный ID игры', 'error')
        return redirect(url_for('add_game'))
    
    add_url = f"https://boardgamegeek.com/xmlapi/boardgame/{game_id}?stats=1"
    response = requests.get(add_url, headers=BGG_HEADERS)
    
    if response.status_code == 200:
        root = ET.fromstring(response.text)
        game = root.find(".//boardgame")
        
        if game is not None:
            names = game.findall("name")
            russian_name = None
            for name in names:
                if contains_cyrillic(name.text):
                    russian_name = name.text
                    break
            if russian_name is None and names:
                russian_name = names[0].text
            
            objectid = game.get("objectid")
            minplayers = game.find("minplayers").text if game.find("minplayers") is not None else ""
            maxplayers = game.find("maxplayers").text if game.find("maxplayers") is not None else ""
            playingtime = game.find("playingtime").text if game.find("playingtime") is not None else ""
            average = game.find(".//average").text if game.find(".//average") is not None else ""
            
            categories = game.findall(".//boardgamecategory")
            categories_list = [category.text for category in categories]
            categories_str = "; ".join(categories_list)
            
            thumbnail = game.find("thumbnail")
            thumbnail_url = thumbnail.text if thumbnail is not None else None
            
            image = game.find("image")
            image_url = image.text if image is not None else None
            
            # Полноразмерная обложка нужна карточкам сайта. Thumbnail имеет
            # высоту около 150 px и заметно размывается при растягивании.
            img_url_to_use = thumbnail_url if thumbnail_url else image_url
            
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            
            next_row = sheet.max_row + 1
            
            sheet.append([russian_name, objectid, minplayers, maxplayers, playingtime, average, categories_str, 'web_user'])
            
            if img_url_to_use:
                download_and_insert_image(sheet, next_row, img_url_to_use, russian_name, objectid)
                
            workbook.save(EXCEL_FILE)
            
            flash(f"Игра '{russian_name}' успешно добавлена!", 'success')
            return redirect(url_for('games'))
    
    flash('Ошибка при добавлении игры', 'error')
    return redirect(url_for('add_game'))


@app.route('/filter')
def filter_games():
    return render_template('filter.html')


@app.route('/filter/players', methods=['GET', 'POST'])
def filter_by_players():
    if request.method == 'POST':
        num_players = request.form.get('num_players')
        if num_players and num_players.isdigit():
            num_players = int(num_players)
            matching_games = []
            for game in get_games_list():
                min_players = int(game['min_players']) if game['min_players'] else 0
                max_players = int(game['max_players']) if game['max_players'] else 0
                if min_players <= num_players <= max_players:
                    matching_games.append(game)
            
            return render_template('filter_results.html', games=matching_games, 
                                 filter_text=f"для {num_players} игроков")
    
    return render_template('filter_players.html')


@app.route('/filter/category', methods=['GET', 'POST'])
def filter_by_category():
    categories = get_all_categories()
    
    if request.method == 'POST':
        category = request.form.get('category')
        if category:
            matching_games = []
            for game in get_games_list():
                if game['categories']:
                    game_categories = game['categories'].split("; ")
                    if category in game_categories:
                        matching_games.append(game)
            
            return render_template('filter_results.html', games=matching_games, 
                                 filter_text=f"в категории '{category}'")
    
    return render_template('filter_category.html', categories=categories)


@app.route('/filter/both', methods=['GET', 'POST'])
def filter_by_both():
    categories = get_all_categories()
    
    if request.method == 'POST':
        category = request.form.get('category')
        num_players = request.form.get('num_players')
        
        if category and num_players and num_players.isdigit():
            num_players = int(num_players)
            matching_games = []
            for game in get_games_list():
                min_players = int(game['min_players']) if game['min_players'] else 0
                max_players = int(game['max_players']) if game['max_players'] else 0
                categories_game = game['categories'].split("; ") if game['categories'] else []
                
                if category in categories_game and min_players <= num_players <= max_players:
                    matching_games.append(game)
            
            return render_template('filter_results.html', games=matching_games,
                                 filter_text=f"в категории '{category}' для {num_players} игроков")
    
    return render_template('filter_both.html', categories=categories)


if __name__ == '__main__':
    app.run(
        host='0.0.0.0',
        port=5000, 
        debug=True
    )

